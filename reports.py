"""
reports.py — Hisobotlarni Excel, PDF va Grafik ko'rinishida yaratish
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import matplotlib
matplotlib.use("Agg")  # ekran kerak emas, faylga chizadi
import matplotlib.pyplot as plt

OUTPUT_DIR = "reports_output"
os.makedirs(OUTPUT_DIR, exist_ok=True)


def generate_excel(report, period_label, filepath=None):
    """Sotuvlar va chiqimlarni Excel faylga yozadi."""
    if filepath is None:
        filepath = os.path.join(OUTPUT_DIR, "hisobot.xlsx")

    wb = Workbook()

    # --- Sotuvlar varag'i ---
    ws1 = wb.active
    ws1.title = "Sotuvlar"
    headers = ["Sana", "Mahsulot", "Miqdor", "Sotish narxi", "Tannarx", "Jami", "Foyda"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for s in report["sales"]:
        ws1.append([
            s["created_at"][:16].replace("T", " "),
            s["item_name"],
            s["quantity"],
            s["sale_price"],
            s["purchase_price"],
            s["total"],
            s["profit"],
        ])

    for col in ws1.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws1.column_dimensions[col[0].column_letter].width = max_len + 3

    # --- Chiqimlar varag'i ---
    ws2 = wb.create_sheet("Chiqimlar")
    ws2.append(["Sana", "Izoh", "Summa"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    for e in report["expenses"]:
        ws2.append([e["created_at"][:16].replace("T", " "), e["description"], e["amount"]])

    for col in ws2.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws2.column_dimensions[col[0].column_letter].width = max_len + 3

    # --- Xulosa varag'i ---
    ws3 = wb.create_sheet("Xulosa")
    ws3.append(["Ko'rsatkich", "Summa (so'm)"])
    ws3.append(["Davr", period_label])
    ws3.append(["Kirim (sotuvlar)", report["total_income"]])
    ws3.append(["Sotuvlardan sof foyda", report["total_gross_profit"]])
    ws3.append(["Chiqimlar", report["total_expenses"]])
    ws3.append(["Sof foyda/zarar", report["net_profit"]])
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    wb.save(filepath)
    return filepath


def generate_pdf(report, period_label, filepath=None):
    """Hisobotni PDF fayl sifatida yaratadi."""
    if filepath is None:
        filepath = os.path.join(OUTPUT_DIR, "hisobot.pdf")

    doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"<b>Do'kon hisoboti</b> — {period_label}", styles["Title"]))
    elements.append(Spacer(1, 12))

    result_label = "SOF FOYDA" if report["net_profit"] >= 0 else "SOF ZARAR"
    summary_data = [
        ["Kirim (sotuvlar)", f"{report['total_income']:,.0f} so'm"],
        ["Chiqimlar", f"{report['total_expenses']:,.0f} so'm"],
        ["Sotuvlardan sof foyda", f"{report['total_gross_profit']:,.0f} so'm"],
        [result_label, f"{report['net_profit']:,.0f} so'm"],
    ]
    summary_table = Table(summary_data, colWidths=[8 * cm, 6 * cm])
    summary_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    if report["sales"]:
        elements.append(Paragraph("<b>Sotuvlar</b>", styles["Heading2"]))
        sale_data = [["Sana", "Mahsulot", "Miqdor", "Narx", "Jami", "Foyda"]]
        for s in report["sales"]:
            sale_data.append([
                s["created_at"][:16].replace("T", " "),
                s["item_name"],
                str(s["quantity"]),
                f"{s['sale_price']:,.0f}",
                f"{s['total']:,.0f}",
                f"{s['profit']:,.0f}",
            ])
        sale_table = Table(sale_data, repeatRows=1)
        sale_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        elements.append(sale_table)
        elements.append(Spacer(1, 16))

    if report["expenses"]:
        elements.append(Paragraph("<b>Chiqimlar</b>", styles["Heading2"]))
        exp_data = [["Sana", "Izoh", "Summa"]]
        for e in report["expenses"]:
            exp_data.append([e["created_at"][:16].replace("T", " "), e["description"], f"{e['amount']:,.0f}"])
        exp_table = Table(exp_data, repeatRows=1)
        exp_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C00000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        elements.append(exp_table)

    doc.build(elements)
    return filepath


def generate_items_pdf(items, filepath=None):
    """Ombordagi mahsulotlar ro'yxatini PDF fayl sifatida yaratadi."""
    if filepath is None:
        filepath = os.path.join(OUTPUT_DIR, "ombor.pdf")

    doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>Ombordagi mahsulotlar</b>", styles["Title"]))
    elements.append(Spacer(1, 12))

    data = [["Mahsulot", "Qoldiq", "Birlik", "Tannarx (so'm)", "Jami qiymat (so'm)"]]
    total_value = 0
    for item in items:
        value = item["quantity"] * item["purchase_price"]
        total_value += value
        data.append([
            item["name"],
            f"{item['quantity']:,.0f}".replace(",", " "),
            item["unit"],
            f"{item['purchase_price']:,.0f}".replace(",", " "),
            f"{value:,.0f}".replace(",", " "),
        ])
    data.append(["", "", "", "Jami:", f"{total_value:,.0f}".replace(",", " ")])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(table)

    doc.build(elements)
    return filepath


def generate_chart(report, period_label, filepath=None):
    """Kunlar bo'yicha foyda/zarar ustunli diagrammasini yaratadi."""
    if filepath is None:
        filepath = os.path.join(OUTPUT_DIR, "grafik.png")

    daily_profit = {}
    for s in report["sales"]:
        day = s["created_at"][:10]
        daily_profit[day] = daily_profit.get(day, 0) + s["profit"]
    for e in report["expenses"]:
        day = e["created_at"][:10]
        daily_profit[day] = daily_profit.get(day, 0) - e["amount"]

    days = sorted(daily_profit.keys())
    values = [daily_profit[d] for d in days]
    bar_colors = ["#2ca02c" if v >= 0 else "#d62728" for v in values]

    plt.figure(figsize=(8, 4.5))
    plt.bar(days, values, color=bar_colors)
    plt.title(f"Kunlik foyda/zarar — {period_label}")
    plt.ylabel("so'm")
    plt.xticks(rotation=45, ha="right", fontsize=8)
    plt.axhline(0, color="black", linewidth=0.8)
    plt.tight_layout()
    plt.savefig(filepath, dpi=150)
    plt.close()
    return filepath
